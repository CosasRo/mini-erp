"""
Orçado x Realizado - Mini ERP PagoExpress
Duas abas: comparação mensal Orçado x Realizado por Centro de Custo (só
visualização), e Compromissos Diários (cadastro manual: incluir, editar,
excluir, filtrar e baixar).
"""

import io
import pandas as pd
import streamlit as st
from supabase import create_client

CENTROS_CUSTO = [
    "13º Salário", "Ações Trab. Hon./Indenização/Custas/Rescisão", "Acordos Jurídicos 2022",
    "Acordos Jurídicos 2022 Extra", "Acordos Jurídicos 2024", "Adiantamento a Fornecedores", "Água e esgoto",
    "Aluguel de imóveis", "Aluguel equipamentos", "Anúncios", "Aquisição de Softwares",
    "Assessoria de Imprensa", "Assessoria e Consultoria", "Assistência Jurídica", "Auditoria",
    "Benfeitoria de Imóveis", "Bens de diminuto valor", "Brindes e Presentes",
    "Cartão de Crédito Corporativo", "Cartório", "Combustíveis e lubrificantes", "Comissões",
    "Condomínio", "Conduções/Táxi", "Conservação de móveis e objetos de decoração",
    "Consultoria - Prestadores de Serviços Internos", "Consultoria de Projetos", "Contabilidade",
    "Contingência", "Contribuição Assistencial", "Contribuição Sindical Patronal-Empresa",
    "Contribuições Sindicais-Empregados", "Copa e Cozinha", "Cópias e Reproduções", "CSLL",
    "Criação da Campanha", "Cursos e Treinamentos", "Custos Adicionais",
    "Despesas com financiamentos", "Despesas de viagens", "Dissídio", "Diversos Marketing",
    "Documentos legais", "Emolumentos e taxas", "Encargos Sociais FGTS", "Encargos Sociais INSS",
    "Energia Elétrica", "Equipamentos de Processamento de Dados", "Estacionamento", "Férias",
    "Ferramentas", "Eventos", "Fundo Fixo de Caixa", "Gráfica", "Gratificação Bonus",
    "Guarda de Documentos", "Higiene e Limpeza", "Imóveis", "Indenizações",
    "Infraestrutura Digital (Site, Softwares Próprios)", "Infrastrutura de Hardware e Telefonia",
    "Instalações", "IOF", "IPTU", "IRRF sobre salários - cód. 0561",
    "Juros - Caixa Economica Federal", "Lanches, refeições", "Livros, Jornais, revistas e TV",
    "Mala Direta", "Manutenção de equipamentos", "Manutenção de Hardware/Software - Informática",
    "COFINS", "Manutenção Predial", "Máquinas aparelhos e Equipamentos",
    "Marcas Direitos e Patentes", "Marketing Institucional", "Material de escritório",
    "Material de Limpeza", "Motoboy", "Móveis e Utensílios", "Multa de FGTS rescisão",
    "Outros Impostos e Taxas", "Parque Gráfico (Tonners e Papéis)", "Pedágio",
    "Pesquisa de Mercado", "Plano de Saude", "Pró Labore", "Programa de alimentação",
    "Projeto de Arquitetura",
    "Distribuição de Dividendos", "IRPJ", "Reembolso diversos", "Salários e Ordenados",
    "Saude Ocupacional", "Segurança", "Seguro de imóvel", "PIS", "Seguro de vida",
    "Serviços de Terceiros", "Serviços de Terceiros - Personalização",
    "Serviços de Terceiros - PJ", "Site", "Softwares", "Tarifas Bancárias",
    "Tarifas Bancárias Extras", "Telefone fixo/fax/internet", "Telefonia Móvel", "Terrenos",
    "Trabalhista", "ISS", "Uniformes", "Vale Transporte/Fretado", "Veículos",
]

BANCOS = ["Banco do Brasil", "Bradesco", "Asaas"]


@st.cache_resource
def get_supabase():
    return create_client(st.secrets["supabase"]["url"], st.secrets["supabase"]["key"])


def _primeiro_dia_mes(data) -> str:
    return pd.Timestamp(data).replace(day=1).strftime("%Y-%m-%d")


def _carregar_orcado(_sb, mes: str) -> pd.DataFrame:
    resp = _sb.table("orcamento").select("centro_custo,valor_orcado").eq("mes", mes).execute()
    df = pd.DataFrame(resp.data)
    if df.empty:
        return pd.DataFrame(columns=["centro_custo", "valor_orcado"])
    df["valor_orcado"] = pd.to_numeric(df["valor_orcado"], errors="coerce").fillna(0.0)
    return df


def _carregar_realizado(_sb, mes: str) -> pd.DataFrame:
    resp = _sb.table("compromissos_diarios").select("centro_custo,valor").eq("mes", mes).execute()
    df = pd.DataFrame(resp.data)
    if df.empty:
        return pd.DataFrame(columns=["centro_custo", "valor"])
    df["valor"] = pd.to_numeric(df["valor"], errors="coerce").fillna(0.0)
    return df.groupby("centro_custo", as_index=False)["valor"].sum()


MESES_PT = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]


def _carregar_orcado_ano(_sb, ano: int) -> pd.DataFrame:
    resp = (
        _sb.table("orcamento")
        .select("centro_custo,mes,valor_orcado")
        .gte("mes", f"{ano}-01-01")
        .lte("mes", f"{ano}-12-31")
        .execute()
    )
    df = pd.DataFrame(resp.data)
    if df.empty:
        return pd.DataFrame(columns=["centro_custo", "mes", "valor_orcado"])
    df["mes"] = pd.to_datetime(df["mes"])
    df["valor_orcado"] = pd.to_numeric(df["valor_orcado"], errors="coerce").fillna(0.0)
    return df


def _carregar_realizado_ano(_sb, ano: int) -> pd.DataFrame:
    resp = (
        _sb.table("compromissos_diarios")
        .select("centro_custo,mes,valor")
        .gte("mes", f"{ano}-01-01")
        .lte("mes", f"{ano}-12-31")
        .execute()
    )
    df = pd.DataFrame(resp.data)
    if df.empty:
        return pd.DataFrame(columns=["centro_custo", "mes", "valor"])
    df["mes"] = pd.to_datetime(df["mes"])
    df["valor"] = pd.to_numeric(df["valor"], errors="coerce").fillna(0.0)
    return df.groupby(["centro_custo", "mes"], as_index=False)["valor"].sum()


HIERARQUIA = {
    "Pessoal": [
        "Pró Labore", "Salários e Ordenados", "Dissídio", "Férias", "Vale Transporte/Fretado",
        "Programa de alimentação", "13 Salário", "Ações Trab. Hon./Indenização/Custas/Rescisão",
        "Plano de Saude", "Encargos Sociais INSS", "Encargos Sociais FGTS", "Gratificação Bonus",
        "Multa de FGTS rescisão", "Contribuições Sindicais-Empregados", "Cursos e Treinamentos",
        "Consultoria - Prestadores de Serviços Internos", "Contribuição Assistencial",
        "IRRF sobre salários - cód. 0561", "Saude Ocupacional", "Trabalhista", "Uniformes",
        "Seguro de vida", "Distribuição de Dividendos", "Contingência",
    ],
    "Ocupação": [
        "Água e esgoto", "Aluguel de imóveis", "Condomínio",
        "Conservação de móveis e objetos de decoração", "Energia Elétrica", "IPTU",
        "Manutenção Predial", "Segurança", "Seguro de imóvel",
    ],
    "Operação do escritório": [
        "Bens de diminuto valor", "Contribuição Sindical Patronal-Empresa", "Copa e Cozinha",
        "Cópias e Reproduções", "Emolumentos e taxas", "Higiene e Limpeza",
        "Livros, Jornais, revistas e TV", "Material de escritório", "Material de Limpeza",
        "Parque Gráfico (Tonners e Papéis)", "Telefone fixo/fax/internet", "Telefonia Móvel",
    ],
    "Gastos gerais": [
        "Adiantamento a Fornecedores", "Aluguel equipamentos", "Cartão de Crédito Corporativo",
        "Combustíveis e lubrificantes", "Conduções/Táxi", "Despesas de viagens", "Estacionamento",
        "Fundo Fixo de Caixa", "Lanches, refeições", "Manutenção de equipamentos",
        "Reembolso diversos", "Outros Impostos e Taxas", "Comissões", "Pedágio", "COFINS",
        "CSLL", "IRPJ", "PIS", "ISS",
    ],
    "Marketing": [
        "Anúncios", "Assessoria de Imprensa", "Brindes e Presentes", "Criação da Campanha",
        "Diversos Marketing", "Eventos", "Gráfica", "Mala Direta", "Marketing Institucional", "Site",
    ],
    "Serviços prof. Contratados": [
        "Assessoria e Consultoria", "Assistência Jurídica", "Acordos Jurídicos 2024",
        "Acordos Jurídicos 2022 Extra", "Auditoria", "Cartório",
        "Consultoria de Projetos", "Contabilidade", "Indenizações", "Custos Adicionais",
        "Despesas com financiamentos", "Documentos legais", "Guarda de Documentos", "Motoboy",
        "Pesquisa de Mercado", "Projeto de Arquitetura", "Serviços de Terceiros",
        "Serviços de Terceiros - Personalização", "Serviços de Terceiros - PJ",
    ],
    "Despesas bancárias": [
        "Tarifas Bancárias Extras", "IOF", "Juros - Caixa Economica Federal", "Tarifas Bancárias",
    ],
}

# Investimentos tem um nível a mais: 2 subgrupos, cada um com suas próprias categorias
SUBGRUPOS_INVESTIMENTOS = {
    "Tecnologia da informação": [
        "Aquisição de Softwares", "Infraestrutura Digital (Site, Softwares Próprios)",
        "Infrastrutura de Hardware e Telefonia", "Manutenção de Hardware/Software - Informática",
        "Softwares",
    ],
    "Ativo Imobilizado": [
        "Benfeitoria de Imóveis", "Equipamentos de Processamento de Dados", "Ferramentas",
        "Imóveis", "Instalações", "Máquinas aparelhos e Equipamentos", "Marcas Direitos e Patentes",
        "Móveis e Utensílios", "Terrenos", "Veículos",
    ],
}


def _gerar_excel_orcado(tabela: pd.DataFrame, eh_grupo: list, meses: list, cor_celula) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orçado x Realizado"

    fill_grupo = PatternFill(start_color="2A2140", end_color="2A2140", fill_type="solid")
    fonte_grupo = Font(bold=True, color="FFFFFF")
    borda_fina = Side(style="thin", color="444444")
    borda_padrao = Border(left=borda_fina, right=borda_fina, top=borda_fina, bottom=borda_fina)

    # cabeçalho: linha 1 = grupo de mês, linha 2 = Orçado/%/Realizado
    for c, (grupo, nome) in enumerate(tabela.columns, start=1):
        ws.cell(row=1, column=c, value=grupo if nome != "Centro de Custo" else "")
        ws.cell(row=2, column=c, value=nome if nome != "Centro de Custo" else "Centro de Custo")
        ws.cell(row=1, column=c).font = Font(bold=True)
        ws.cell(row=2, column=c).font = Font(bold=True)
        ws.cell(row=1, column=c).border = borda_padrao
        ws.cell(row=2, column=c).border = borda_padrao

    for r, (_, linha) in enumerate(tabela.iterrows(), start=3):
        eh = eh_grupo[r - 3]
        for c, (grupo, nome) in enumerate(tabela.columns, start=1):
            valor = linha[(grupo, nome)]
            cel = ws.cell(row=r, column=c, value=valor)
            cel.border = borda_padrao
            if nome == "%":
                cel.number_format = "0%"
            elif nome in ("Orçado", "Realizado", "Previsto"):
                cel.number_format = "#,##0.00"

            cor_hex = cor_celula(nome, valor) if nome != "Centro de Custo" else None
            if eh:
                cel.fill = fill_grupo
                cel.font = fonte_grupo if not cor_hex else Font(bold=True, color=cor_hex.lstrip("#"))
            elif cor_hex:
                cel.font = Font(color=cor_hex.lstrip("#"))

    for c in range(1, len(tabela.columns) + 1):
        ws.column_dimensions[ws.cell(row=2, column=c).column_letter].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _gerar_pdf_orcado(tabela: pd.DataFrame, eh_grupo: list, meses: list, cor_celula) -> bytes:
    from reportlab.lib.pagesizes import landscape, A3
    from reportlab.lib import colors as rl_colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A3), leftMargin=15, rightMargin=15, topMargin=20, bottomMargin=15)
    styles = getSampleStyleSheet()
    elementos = [Paragraph("<b>Pago Express — Orçado x Realizado</b>", styles["Title"]), Spacer(1, 8)]

    cabecalho1 = [g if n != "Centro de Custo" else "" for g, n in tabela.columns]
    cabecalho2 = [n if n != "Centro de Custo" else "Centro de Custo" for g, n in tabela.columns]
    dados = [cabecalho1, cabecalho2]

    for _, linha in tabela.iterrows():
        dados.append([
            f"{linha[(g, n)]:.0%}" if n == "%" else
            (f"{linha[(g, n)]:,.0f}" if n in ("Orçado", "Realizado", "Previsto") else str(linha[(g, n)]))
            for g, n in tabela.columns
        ])

    tbl = Table(dados, repeatRows=2)
    estilo = [
        ("BACKGROUND", (0, 0), (-1, 1), rl_colors.HexColor("#6c3fa8")),
        ("TEXTCOLOR", (0, 0), (-1, 1), rl_colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 5),
        ("GRID", (0, 0), (-1, -1), 0.4, rl_colors.grey),
    ]
    for r, eh in enumerate(eh_grupo, start=2):
        if eh:
            estilo.append(("BACKGROUND", (0, r), (-1, r), rl_colors.HexColor("#2a2140")))
            estilo.append(("TEXTCOLOR", (0, r), (-1, r), rl_colors.white))
            estilo.append(("FONTNAME", (0, r), (-1, r), "Helvetica-Bold"))
        for c, (g, n) in enumerate(tabela.columns):
            if n == "Centro de Custo":
                continue
            valor = tabela.iloc[r - 2][(g, n)]
            cor_hex = cor_celula(n, valor)
            if cor_hex and not eh:
                estilo.append(("TEXTCOLOR", (c, r), (c, r), rl_colors.HexColor(cor_hex)))
    tbl.setStyle(TableStyle(estilo))
    elementos.append(tbl)
    doc.build(elementos)
    return buf.getvalue()


def render_orcado_realizado():
    sb = get_supabase()

    ano = st.selectbox("Ano", [2025, 2026, 2027], index=1)

    orcado = _carregar_orcado_ano(sb, ano)
    realizado = _carregar_realizado_ano(sb, ano)

    if orcado.empty and realizado.empty:
        st.info(f"Nenhum dado de orçamento ou compromissos para {ano} ainda.")
        return

    orc_por_centro_mes = orcado.set_index(["centro_custo", orcado["mes"].dt.month])["valor_orcado"]
    rea_por_centro_mes = realizado.set_index(["centro_custo", realizado["mes"].dt.month])["valor"]

    def _valor(serie, centro, mes_num):
        try:
            return float(serie.loc[(centro, mes_num)])
        except KeyError:
            return 0.0

    def _linha_leaf(nome):
        orc = [_valor(orc_por_centro_mes, nome, m) for m in range(1, 13)]
        rea = [_valor(rea_por_centro_mes, nome, m) for m in range(1, 13)]
        return orc, rea

    # linhas: (label, nivel, orc[12], rea[12]) -- nivel 0=grupo, 1=subgrupo, 2=folha
    linhas = []
    orc_geral = [0.0] * 12
    rea_geral = [0.0] * 12

    for grupo, leaves in HIERARQUIA.items():
        orc_grupo = [0.0] * 12
        rea_grupo = [0.0] * 12
        leaf_rows = []
        for leaf in leaves:
            orc_leaf, rea_leaf = _linha_leaf(leaf)
            leaf_rows.append((leaf, 2, orc_leaf, rea_leaf))
            for i in range(12):
                orc_grupo[i] += orc_leaf[i]
                rea_grupo[i] += rea_leaf[i]
        linhas.append(("␀SPACER␀", None, None, None))
        linhas.append((grupo.upper(), 0, orc_grupo, rea_grupo))
        linhas.extend(leaf_rows)
        for i in range(12):
            orc_geral[i] += orc_grupo[i]
            rea_geral[i] += rea_grupo[i]

    # Investimentos -- 3 níveis: grupo > subgrupo > folha
    orc_investimentos = [0.0] * 12
    rea_investimentos = [0.0] * 12
    linhas_subgrupos = []
    for subgrupo, leaves in SUBGRUPOS_INVESTIMENTOS.items():
        orc_sub = [0.0] * 12
        rea_sub = [0.0] * 12
        leaf_rows = []
        for leaf in leaves:
            orc_leaf, rea_leaf = _linha_leaf(leaf)
            leaf_rows.append((leaf, 2, orc_leaf, rea_leaf))
            for i in range(12):
                orc_sub[i] += orc_leaf[i]
                rea_sub[i] += rea_leaf[i]
        linhas_subgrupos.append((subgrupo, 1, orc_sub, rea_sub))
        linhas_subgrupos.extend(leaf_rows)
        for i in range(12):
            orc_investimentos[i] += orc_sub[i]
            rea_investimentos[i] += rea_sub[i]

    linhas.append(("␀SPACER␀", None, None, None))
    linhas.append(("INVESTIMENTOS", 0, orc_investimentos, rea_investimentos))
    linhas.extend(linhas_subgrupos)
    for i in range(12):
        orc_geral[i] += orc_investimentos[i]
        rea_geral[i] += rea_investimentos[i]

    linhas.append(("␀SPACER␀", None, None, None))
    linhas.append(("DESPESAS OPERACIONAIS (TOTAL)", 0, orc_geral, rea_geral))

    # ------------------------------------------------------------------
    # Monta a tabela HTML (dá controle real de borda/altura, diferente do
    # st.dataframe que não desenha CSS de borda).
    # ------------------------------------------------------------------
    COR_ORCADO, COR_REALIZADO = "#f2a5a5", "#9cc3f2"
    COR_PCT_OK, COR_PCT_ESTOURO = "#8fd6a5", "#f2a5a5"
    COR_DIVISOR = "#5a4d80"

    def _fmt_moeda(v):
        return f"R$ {v:,.0f}"

    cabecalho_meses = "".join(
        f'<th colspan="3" style="padding:6px 4px; text-align:center; color:#fff; '
        f'border-left:2px solid {COR_DIVISOR};">{m}</th>'
        for m in MESES_PT
    )
    cabecalho_sub = "".join(
        f'<th style="padding:4px; text-align:right; color:#c9c7c1; font-weight:400; '
        f'border-left:2px solid {COR_DIVISOR};">Orçado</th>'
        f'<th style="padding:4px; text-align:right; color:#c9c7c1; font-weight:400;">%</th>'
        f'<th style="padding:4px; text-align:right; color:#c9c7c1; font-weight:400;">Realizado</th>'
        for _ in MESES_PT
    )

    CATEGORIAS_FORA_DO_GERAL = {
        "Distribuição de Dividendos", "Comissões", "COFINS", "CSLL", "IRPJ", "PIS", "ISS",
    }

    def _linha_html(label, orc, rea, estilo_linha, recuo="0px"):
        celulas = [f'<td style="padding:4px 6px; {estilo_linha} padding-left:{recuo};">{label}</td>']
        for m_idx in range(12):
            o, r = orc[m_idx], rea[m_idx]
            pct = (r / o) if o else 0.0
            cor_pct = COR_PCT_ESTOURO if pct > 1.0 else COR_PCT_OK
            celulas.append(
                f'<td style="padding:4px; text-align:right; {estilo_linha} color:{COR_ORCADO}; '
                f'border-left:2px solid {COR_DIVISOR};">{_fmt_moeda(o)}</td>'
                f'<td style="padding:4px; text-align:right; {estilo_linha} color:{cor_pct};">{pct*100:.0f}%</td>'
                f'<td style="padding:4px; text-align:right; {estilo_linha} color:{COR_REALIZADO};">{_fmt_moeda(r)}</td>'
            )
        total_orc, total_rea = sum(orc), sum(rea)
        pct_total = (total_rea / total_orc) if total_orc else 0.0
        cor_pct_total = COR_PCT_ESTOURO if pct_total > 1.0 else COR_PCT_OK
        celulas.append(
            f'<td style="padding:4px; text-align:right; {estilo_linha} color:{COR_ORCADO}; '
            f'border-left:2px solid {COR_DIVISOR};">{_fmt_moeda(total_orc)}</td>'
            f'<td style="padding:4px; text-align:right; {estilo_linha} color:{COR_REALIZADO};">{_fmt_moeda(total_rea)}</td>'
            f'<td style="padding:4px; text-align:right; {estilo_linha} color:{cor_pct_total};">{pct_total*100:.0f}%</td>'
        )
        return f'<tr>{"".join(celulas)}</tr>'

    # total ajustado -- mesma soma geral, mas sem as categorias que não entram
    # no total de gastos gerais (ex: dividendos, comissões, impostos sobre lucro)
    orc_ajustado = [0.0] * 12
    rea_ajustado = [0.0] * 12
    for label, nivel, orc, rea in linhas:
        if nivel == 2 and label not in CATEGORIAS_FORA_DO_GERAL:
            for i in range(12):
                orc_ajustado[i] += orc[i]
                rea_ajustado[i] += rea[i]

    linha_total_geral_html = _linha_html(
        "TOTAL GERAL", orc_ajustado, rea_ajustado, "font-weight:800; background:#3A2D58; color:#fff;"
    )
    linha_total_ajustado_html = _linha_html(
        "TOTAL GERAL (c/ dividendos, comissões e impostos sobre lucro)",
        orc_geral, rea_geral, "font-weight:800; background:#4a3d6b; color:#fff;"
    )

    linhas_html = []
    for label, nivel, orc, rea in linhas:
        if label == "␀SPACER␀":
            linhas_html.append(f'<tr style="height:5px; background:{COR_DIVISOR}77;">'
                                f'<td colspan="{1 + 12*3 + 3}" style="padding:0;"></td></tr>')
            continue

        if nivel == 0:
            estilo_linha = "font-weight:700; background:#2a2140; color:#fff;"
            recuo = "0px"
        elif nivel == 1:
            estilo_linha = "font-weight:600; background:#231c37; color:#e6e1f5;"
            recuo = "14px"
        else:
            estilo_linha = "color:#e6e1f5;"
            recuo = "28px"

        celulas = [f'<td style="padding:4px 6px; {estilo_linha} padding-left:{recuo};">{label}</td>']
        for m_idx in range(12):
            o, r = orc[m_idx], rea[m_idx]
            pct = (r / o) if o else 0.0
            cor_pct = COR_PCT_ESTOURO if pct > 1.0 else COR_PCT_OK
            celulas.append(
                f'<td style="padding:4px; text-align:right; {estilo_linha} color:{COR_ORCADO}; '
                f'border-left:2px solid {COR_DIVISOR};">{_fmt_moeda(o)}</td>'
                f'<td style="padding:4px; text-align:right; {estilo_linha} color:{cor_pct};">{pct*100:.0f}%</td>'
                f'<td style="padding:4px; text-align:right; {estilo_linha} color:{COR_REALIZADO};">{_fmt_moeda(r)}</td>'
            )
        total_orc, total_rea = sum(orc), sum(rea)
        pct_total = (total_rea / total_orc) if total_orc else 0.0
        cor_pct_total = COR_PCT_ESTOURO if pct_total > 1.0 else COR_PCT_OK
        celulas.append(
            f'<td style="padding:4px; text-align:right; {estilo_linha} color:{COR_ORCADO}; '
            f'border-left:2px solid {COR_DIVISOR};">{_fmt_moeda(total_orc)}</td>'
            f'<td style="padding:4px; text-align:right; {estilo_linha} color:{COR_REALIZADO};">{_fmt_moeda(total_rea)}</td>'
            f'<td style="padding:4px; text-align:right; {estilo_linha} color:{cor_pct_total};">{pct_total*100:.0f}%</td>'
        )
        linhas_html.append(f'<tr>{"".join(celulas)}</tr>')

    html = f"""
    <div style="overflow-x:auto; max-height:650px; overflow-y:auto;">
    <table style="width:100%; border-collapse:collapse; font-size:12px; white-space:nowrap;">
        <thead style="position:sticky; top:0; background:#3A2D58; z-index:1;">
            <tr>
                <th rowspan="2" style="padding:6px; text-align:left; color:#fff;">Centro de Custo</th>
                {cabecalho_meses}
                <th colspan="3" style="padding:6px 4px; text-align:center; color:#fff;
                    border-left:2px solid {COR_DIVISOR};">Total Ano</th>
            </tr>
            <tr>
                {cabecalho_sub}
                <th style="padding:4px; text-align:right; color:#c9c7c1; font-weight:400;
                    border-left:2px solid {COR_DIVISOR};">Previsto</th>
                <th style="padding:4px; text-align:right; color:#c9c7c1; font-weight:400;">Realizado</th>
                <th style="padding:4px; text-align:right; color:#c9c7c1; font-weight:400;">%</th>
            </tr>
        </thead>
        <tbody>
            {linha_total_geral_html}
            {linha_total_ajustado_html}
            <tr style="height:8px;"><td colspan="{1 + 12*3 + 3}" style="padding:0;"></td></tr>
            {"".join(linhas_html)}
        </tbody>
    </table>
    </div>
    """
    st.markdown(html, unsafe_allow_html=True)

    # ------------------------------------------------------------------
    # Excel/PDF -- reutiliza as mesmas funções de antes, com um DataFrame
    # "achatado" (sem os espaçadores) gerado a partir das mesmas linhas.
    # ------------------------------------------------------------------
    linhas_export = [l for l in linhas if l[0] != "␀SPACER␀"]
    dados = {}
    for m_idx, m in enumerate(MESES_PT):
        dados[(m, "Orçado")] = [l[2][m_idx] for l in linhas_export]
        dados[(m, "Realizado")] = [l[3][m_idx] for l in linhas_export]
        dados[(m, "%")] = [(l[3][m_idx] / l[2][m_idx]) if l[2][m_idx] else 0.0 for l in linhas_export]
    dados[("Total Ano", "Previsto")] = [sum(l[2]) for l in linhas_export]
    dados[("Total Ano", "Realizado")] = [sum(l[3]) for l in linhas_export]
    dados[("Total Ano", "%")] = [(sum(l[3]) / sum(l[2])) if sum(l[2]) else 0.0 for l in linhas_export]
    dados[("", "Centro de Custo")] = [l[0] for l in linhas_export]

    ordem = [("", "Centro de Custo")]
    ordem += [c for m in MESES_PT for c in [(m, "Orçado"), (m, "%"), (m, "Realizado")]]
    ordem += [("Total Ano", "Previsto"), ("Total Ano", "Realizado"), ("Total Ano", "%")]
    tabela_export = pd.DataFrame(dados)[ordem]
    tabela_export.columns = pd.MultiIndex.from_tuples(ordem)
    eh_grupo_export = [l[1] in (0, 1) for l in linhas_export]

    def _cor_celula(col_nome, valor):
        if col_nome in ("Orçado", "Previsto"):
            return "#f2a5a5"
        if col_nome == "Realizado":
            return "#9cc3f2"
        if col_nome == "%":
            return "#f2a5a5" if valor > 1.0 else "#8fd6a5"
        return None

    st.markdown("<br>", unsafe_allow_html=True)
    col_dl1, col_dl2 = st.columns(2)
    col_dl1.download_button(
        "⬇️ Baixar (Excel)",
        data=_gerar_excel_orcado(tabela_export, eh_grupo_export, MESES_PT, _cor_celula),
        file_name=f"orcado_x_realizado_{ano}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    col_dl2.download_button(
        "⬇️ Baixar (PDF)",
        data=_gerar_pdf_orcado(tabela_export, eh_grupo_export, MESES_PT, _cor_celula),
        file_name=f"orcado_x_realizado_{ano}.pdf",
        mime="application/pdf",
    )


# ---------------------------------------------------------------
# Compromissos Diários -- aqui sim tem filtro, download e CRUD completo
# ---------------------------------------------------------------
COLUNAS_COMPROMISSO = [
    "id", "banco", "data_pagamento", "documento", "parcela", "emissao",
    "agente", "centro_custo", "valor", "impostos", "saldo", "historico", "mes",
]


def _carregar_compromissos(_sb, data_ini, data_fim, centro_sel) -> pd.DataFrame:
    query = (
        _sb.table("compromissos_diarios")
        .select("*")
        .gte("data_pagamento", str(data_ini))
        .lte("data_pagamento", str(data_fim))
        .order("data_pagamento")
    )
    if centro_sel != "Todos":
        query = query.eq("centro_custo", centro_sel)
    resp = query.execute()
    df = pd.DataFrame(resp.data)
    if df.empty:
        df = pd.DataFrame(columns=COLUNAS_COMPROMISSO)
    for col in COLUNAS_COMPROMISSO:
        if col not in df.columns:
            df[col] = None

    df = df[COLUNAS_COMPROMISSO].copy()
    df["id"] = pd.to_numeric(df["id"], errors="coerce").astype("Int64")
    for col in ("data_pagamento", "emissao", "mes"):
        df[col] = pd.to_datetime(df[col], errors="coerce")
    for col in ("valor", "impostos", "saldo"):
        df[col] = pd.to_numeric(df[col], errors="coerce").astype("float64")
    for col in ("banco", "documento", "parcela", "agente", "centro_custo", "historico"):
        df[col] = df[col].astype("object").where(df[col].notna(), None)

    return df


def _gerar_excel(df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.drop(columns=["id"], errors="ignore").to_excel(writer, index=False, sheet_name="Compromissos")
    return buf.getvalue()


def _gerar_pdf(df: pd.DataFrame) -> bytes:
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=20, rightMargin=20, topMargin=30, bottomMargin=20)
    styles = getSampleStyleSheet()
    elementos = [Paragraph("<b>Pago Express — Compromissos Diários</b>", styles["Title"]), Spacer(1, 10)]

    colunas = ["data_pagamento", "banco", "agente", "centro_custo", "valor", "impostos", "saldo", "historico"]
    colunas = [c for c in colunas if c in df.columns]
    dados = [colunas]
    for _, linha in df.iterrows():
        dados.append([
            linha[c].strftime("%d/%m/%Y") if c == "data_pagamento" and pd.notna(linha[c]) else
            (f"{linha[c]:,.2f}" if c in ("valor", "impostos", "saldo") and pd.notna(linha[c]) else str(linha[c] or ""))
            for c in colunas
        ])

    tabela = Table(dados, repeatRows=1)
    tabela.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#6c3fa8")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f2f2f2")]),
    ]))
    elementos.append(tabela)
    doc.build(elementos)
    return buf.getvalue()


def render_compromissos_diarios():
    sb = get_supabase()

    col1, col2, col3 = st.columns(3)
    data_ini = col1.date_input("De", key="cd_ini", value=pd.Timestamp.today().replace(day=1))
    data_fim = col2.date_input("Até", key="cd_fim")
    centro_sel = col3.selectbox("Centro de Custo", ["Todos"] + CENTROS_CUSTO, key="cd_centro")

    df = _carregar_compromissos(sb, data_ini, data_fim, centro_sel)

    col_dl1, col_dl2 = st.columns(2)
    col_dl1.download_button(
        "⬇️ Baixar (Excel)",
        data=_gerar_excel(df),
        file_name="compromissos_diarios.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        disabled=df.empty,
    )
    col_dl2.download_button(
        "⬇️ Baixar (PDF)",
        data=_gerar_pdf(df) if not df.empty else b"",
        file_name="compromissos_diarios.pdf",
        mime="application/pdf",
        disabled=df.empty,
    )

    st.caption("Edite valores direto na tabela, use a última linha para incluir um novo compromisso, ou marque a caixinha e aperte Delete para excluir uma linha. Depois clique em Salvar.")

    editado = st.data_editor(
        df,
        key="editor_compromissos",
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "id": st.column_config.NumberColumn("id", disabled=True),
            "banco": st.column_config.SelectboxColumn("Banco", options=BANCOS),
            "data_pagamento": st.column_config.DateColumn("Data Pagamento"),
            "documento": st.column_config.TextColumn("Documento"),
            "parcela": st.column_config.TextColumn("Parcela"),
            "emissao": st.column_config.DateColumn("Emissão"),
            "agente": st.column_config.TextColumn("Agente"),
            "centro_custo": st.column_config.SelectboxColumn("Centro de Custo", options=CENTROS_CUSTO),
            "valor": st.column_config.NumberColumn("Valor", format="R$ %.2f"),
            "impostos": st.column_config.NumberColumn("Impostos", format="R$ %.2f"),
            "saldo": st.column_config.NumberColumn("Saldo", format="R$ %.2f"),
            "historico": st.column_config.TextColumn("Histórico"),
            "mes": st.column_config.DateColumn("Mês"),
        },
    )

    if st.button("💾 Salvar alterações"):
        ids_originais = set(df["id"].dropna().astype(int)) if not df.empty else set()
        ids_editados = set(editado["id"].dropna().astype(int)) if "id" in editado.columns else set()

        removidos = ids_originais - ids_editados
        for id_remover in removidos:
            sb.table("compromissos_diarios").delete().eq("id", int(id_remover)).execute()

        for _, linha in editado.iterrows():
            registro = linha.to_dict()
            id_linha = registro.pop("id", None)

            for campo in ("data_pagamento", "emissao", "mes"):
                if pd.notna(registro.get(campo)):
                    registro[campo] = pd.Timestamp(registro[campo]).strftime("%Y-%m-%d")
                else:
                    registro[campo] = None

            if not registro.get("centro_custo"):
                continue  # linha em branco, ignora

            if pd.notna(id_linha):
                sb.table("compromissos_diarios").update(registro).eq("id", int(id_linha)).execute()
            else:
                sb.table("compromissos_diarios").insert(registro).execute()

        st.success("Alterações salvas.")
        st.cache_data.clear()
        st.rerun()

    st.caption(f"{len(df)} lançamento(s) — soma: R$ {df['valor'].sum():,.2f}" if not df.empty else "")


def render_orcamento():
    aba_orcado, aba_compromissos = st.tabs(["Orçado e Realizado", "Compromissos Diários"])

    with aba_orcado:
        render_orcado_realizado()

    with aba_compromissos:
        render_compromissos_diarios()


if __name__ == "__main__":
    render_orcamento()
